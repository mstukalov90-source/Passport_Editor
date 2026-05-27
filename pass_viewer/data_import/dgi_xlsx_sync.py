"""
Stream dgi.xlsx and batch-UPDATE / INSERT rows in the PostGIS dgi table (match on descr).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.db import connection, transaction
from openpyxl import load_workbook
from psycopg2.extras import execute_values


@dataclass(frozen=True)
class DgiXlsxSyncStats:
    rows_seen: int = 0
    skipped_empty_descr: int = 0
    duplicate_descr_in_file: int = 0
    updated: int = 0
    inserted: int = 0
    inserted_placeholder_geom: int = 0
    batches: int = 0

    @property
    def no_match(self) -> int:
        """Approximate count of xlsx rows that did not match any existing DB row on UPDATE."""
        return max(
            0,
            self.rows_seen - self.updated - self.inserted - self.duplicate_descr_in_file,
        )


def _quote_ident(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return None
    s = str(value).strip()
    return s or None


def _resolve_column_index(headers: list[str], *candidates: str) -> int | None:
    lower_headers = [_normalize_header(h) for h in headers]
    for candidate in candidates:
        key = candidate.lower()
        if key in lower_headers:
            return lower_headers.index(key)
    return None


def _fetch_table_meta(table_name: str) -> dict[str, Any]:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT a.attname, t.typname, a.attnotnull
            FROM pg_attribute a
            JOIN pg_class c ON a.attrelid = c.oid
            JOIN pg_namespace n ON c.relnamespace = n.oid
            JOIN pg_type t ON a.atttypid = t.oid
            WHERE n.nspname = 'public'
              AND c.relname = %s
              AND a.attnum > 0
              AND NOT a.attisdropped
            ORDER BY a.attnum
            """,
            [table_name],
        )
        cols = cursor.fetchall()
    if not cols:
        raise ValueError(f'Table "{table_name}" not found in schema public.')

    descr_col = None
    geom_cols: list[tuple[str, bool]] = []
    has_short = False
    for name, typname, notnull in cols:
        if name.lower() == "descr":
            descr_col = name
        if typname == "geometry":
            geom_cols.append((name, bool(notnull)))
        if name.lower() == "short_sobstv_rr":
            has_short = True

    if descr_col is None:
        raise ValueError(f'Table "{table_name}" has no descr column.')
    if not has_short:
        raise ValueError(
            f'Table "{table_name}" has no short_sobstv_rr column; run migrations first.'
        )
    if len(geom_cols) != 1:
        names = [c[0] for c in geom_cols]
        raise ValueError(
            f'Table "{table_name}" must have exactly one geometry column for xlsx insert; found {names or "none"}'
        )
    geom_name, geom_notnull = geom_cols[0]
    return {
        "descr_col": descr_col,
        "geom_col": geom_name,
        "geom_notnull": geom_notnull,
    }


def _update_sql(table: str, descr_col: str, sync_attrs: bool) -> str:
    qt = _quote_ident(table)
    qd = _quote_ident(descr_col)
    if sync_attrs:
        return f"""
            UPDATE {qt} AS t
            SET
                short_sobstv_rr = v.short_sobstv_rr::text,
                address = CASE
                    WHEN v.address IS NOT NULL AND btrim(v.address::text) <> ''
                    THEN v.address::text
                    ELSE t.address
                END,
                sobstv_rr = CASE
                    WHEN v.sobstv_rr IS NOT NULL AND btrim(v.sobstv_rr::text) <> ''
                    THEN v.sobstv_rr::text
                    ELSE t.sobstv_rr
                END
            FROM (VALUES %s) AS v(descr, short_sobstv_rr, address, sobstv_rr)
            WHERE btrim(t.{qd}::text) = btrim(v.descr::text)
        """
    return f"""
        UPDATE {qt} AS t
        SET short_sobstv_rr = v.short_sobstv_rr::text
        FROM (VALUES %s) AS v(descr, short_sobstv_rr)
        WHERE btrim(t.{qd}::text) = btrim(v.descr::text)
    """


def _insert_sql(table: str, descr_col: str, geom_col: str, geom_notnull: bool, target_srid: int) -> str:
    qt = _quote_ident(table)
    qd = _quote_ident(descr_col)
    qg = _quote_ident(geom_col)
    if geom_notnull:
        geom_expr = f"ST_SetSRID(ST_GeomFromText('POLYGON EMPTY'), {int(target_srid)})"
    else:
        geom_expr = "NULL::geometry"
    return f"""
        INSERT INTO {qt} (descr, address, sobstv_rr, short_sobstv_rr, {qg})
        SELECT
            v.descr::text,
            NULLIF(btrim(v.address::text), ''),
            NULLIF(btrim(v.sobstv_rr::text), ''),
            v.short_sobstv_rr::text,
            {geom_expr}
        FROM (VALUES %s) AS v(descr, address, sobstv_rr, short_sobstv_rr)
        WHERE NOT EXISTS (
            SELECT 1 FROM {qt} t
            WHERE btrim(t.{qd}::text) = btrim(v.descr::text)
        )
    """


def _row_tuple(
    cells: tuple[Any, ...],
    *,
    idx_descr: int,
    idx_short: int,
    idx_address: int | None,
    idx_sobstv: int | None,
) -> tuple[str, str | None, str | None, str | None] | None:
    """Returns (descr, short_sobstv_rr, address, sobstv_rr)."""
    descr = _as_text(cells[idx_descr] if idx_descr < len(cells) else None)
    if not descr:
        return None
    short = _as_text(cells[idx_short] if idx_short < len(cells) else None)
    address = _as_text(cells[idx_address] if idx_address is not None and idx_address < len(cells) else None)
    sobstv = _as_text(cells[idx_sobstv] if idx_sobstv is not None and idx_sobstv < len(cells) else None)
    return (descr, short, address, sobstv)


def _update_batch_rows(
    rows: list[tuple[str, str | None, str | None, str | None]],
    *,
    sync_attrs: bool,
) -> list[tuple[Any, ...]]:
    if sync_attrs:
        return [(r[0], r[1], r[2], r[3]) for r in rows]
    return [(r[0], r[1]) for r in rows]


def _insert_batch_rows(
    rows: list[tuple[str, str | None, str | None, str | None]],
) -> list[tuple[Any, ...]]:
    return [(r[0], r[2], r[3], r[1]) for r in rows]


def _load_rows_from_xlsx(
    path: Path,
) -> tuple[dict[str, tuple[str, str | None, str | None, str | None]], int, int]:
    """Returns (unique rows by descr, skipped_empty_descr, duplicate_descr_count)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            raise ValueError(f"Empty workbook: {path}") from None

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        idx_descr = _resolve_column_index(headers, "descr")
        idx_short = _resolve_column_index(headers, "short_sobstv_rr", "Short_sobstv_rr")
        idx_address = _resolve_column_index(headers, "address")
        idx_sobstv = _resolve_column_index(headers, "sobstv_rr")
        if idx_descr is None:
            raise ValueError('dgi.xlsx must contain a "descr" column.')
        if idx_short is None:
            raise ValueError('dgi.xlsx must contain a "Short_sobstv_rr" (or short_sobstv_rr) column.')

        unique: dict[str, tuple[Any, ...]] = {}
        skipped = 0
        duplicates = 0
        for cells in row_iter:
            if cells is None or all(c is None for c in cells):
                continue
            row = _row_tuple(
                cells,
                idx_descr=idx_descr,
                idx_short=idx_short,
                idx_address=idx_address,
                idx_sobstv=idx_sobstv,
            )
            if row is None:
                skipped += 1
                continue
            descr_key = row[0]
            if descr_key in unique:
                duplicates += 1
            unique[descr_key] = row
        return unique, skipped, duplicates
    finally:
        wb.close()


def _flush_batches(
    rows: list[tuple[str, str | None, str | None, str | None]],
    *,
    update_sql: str,
    insert_sql: str,
    update_template: str,
    insert_template: str,
    batch_size: int,
    dry_run: bool,
    geom_notnull: bool,
    sync_attrs: bool,
) -> tuple[int, int, int, int]:
    """Returns (updated, inserted, inserted_placeholder_geom, batches)."""
    updated = 0
    inserted = 0
    placeholder = 0
    batches = 0
    for i in range(0, len(rows), batch_size):
        chunk = rows[i : i + batch_size]
        update_chunk = _update_batch_rows(chunk, sync_attrs=sync_attrs)
        insert_chunk = _insert_batch_rows(chunk)
        batches += 1
        if dry_run:
            updated += len(update_chunk)
            inserted += len(insert_chunk)
            if geom_notnull:
                placeholder += len(insert_chunk)
            continue
        with transaction.atomic():
            with connection.cursor() as dj_cursor:
                raw_cursor = getattr(dj_cursor, "cursor", dj_cursor)
                execute_values(
                    raw_cursor,
                    update_sql,
                    update_chunk,
                    template=update_template,
                    page_size=len(update_chunk),
                )
                updated += dj_cursor.rowcount
                execute_values(
                    raw_cursor,
                    insert_sql,
                    insert_chunk,
                    template=insert_template,
                    page_size=len(insert_chunk),
                )
                n_inserted = dj_cursor.rowcount
                inserted += n_inserted
                if geom_notnull and n_inserted:
                    placeholder += n_inserted
    return updated, inserted, placeholder, batches


def sync_dgi_from_xlsx(
    path: str | Path,
    *,
    table_name: str = "dgi",
    dry_run: bool = False,
    batch_size: int = 2000,
    sync_attrs: bool = False,
    target_srid: int = 4326,
) -> DgiXlsxSyncStats:
    path = Path(path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(path)

    if batch_size < 1:
        raise ValueError("batch_size must be >= 1")

    meta = _fetch_table_meta(table_name)
    descr_col = meta["descr_col"]
    geom_notnull = meta["geom_notnull"]

    unique, skipped, duplicates = _load_rows_from_xlsx(path)
    rows = list(unique.values())

    update_sql = _update_sql(table_name, descr_col, sync_attrs)
    insert_sql = _insert_sql(table_name, descr_col, meta["geom_col"], geom_notnull, target_srid)
    update_template = "(%s, %s, %s, %s)" if sync_attrs else "(%s, %s)"
    insert_template = "(%s, %s, %s, %s)"

    updated, inserted, placeholder, batches = _flush_batches(
        rows,
        update_sql=update_sql,
        insert_sql=insert_sql,
        update_template=update_template,
        insert_template=insert_template,
        batch_size=batch_size,
        dry_run=dry_run,
        geom_notnull=geom_notnull,
        sync_attrs=sync_attrs,
    )

    return DgiXlsxSyncStats(
        rows_seen=len(unique),
        skipped_empty_descr=skipped,
        duplicate_descr_in_file=duplicates,
        updated=updated,
        inserted=inserted,
        inserted_placeholder_geom=placeholder,
        batches=batches,
    )
